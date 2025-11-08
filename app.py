from fastapi import FastAPI, Request
from fastapi.middleware.cors import CORSMiddleware
from fastapi.responses import FileResponse, JSONResponse
from fastapi.staticfiles import StaticFiles
from openpyxl import Workbook, load_workbook
from datetime import datetime
from user_agents import parse
import os
import requests

app = FastAPI()

# Enable CORS
app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"],
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)

# Excel file path
EXCEL_FILE = 'visitors_data.xlsx'

def init_excel():
    """Initialize Excel file with headers if it doesn't exist."""
    if not os.path.exists(EXCEL_FILE):
        wb = Workbook()
        ws = wb.active
        ws.title = "Visitors"
        headers = ['Timestamp', 'IP Address', 'Country', 'City', 'Region',
                   'Browser', 'OS', 'Device', 'Referrer', 'Page URL']
        ws.append(headers)
        wb.save(EXCEL_FILE)
        print(f" Created {EXCEL_FILE}")

def get_location_info(ip):
    """Get location data from IP."""
    try:
        if ip in ['127.0.0.1', 'localhost', '::1']:
            return {'country': 'Local', 'city': 'Local', 'region': 'Local'}

        res = requests.get(f'http://ip-api.com/json/{ip}', timeout=3).json()
        if res.get('status') == 'success':
            return {
                'country': res.get('country', 'Unknown'),
                'city': res.get('city', 'Unknown'),
                'region': res.get('regionName', 'Unknown')
            }
    except:
        pass

    return {'country': 'Unknown', 'city': 'Unknown', 'region': 'Unknown'}

# Serve Frontend Static Files
app.mount("/", StaticFiles(directory="website", html=True), name="website")

@app.post("/track")
async def track_visitor(request: Request):
    try:
        # IP extraction
        ip = request.headers.get('X-Forwarded-For', request.client.host).split(',')[0]
        print(f" New visitor from IP: {ip}")

        # Location info
        location = get_location_info(ip)

        # User Agent
        user_agent_string = request.headers.get("User-Agent", "")
        user_agent = parse(user_agent_string)

        data = await request.json()

        visitor_data = [
            datetime.now().strftime('%Y-%m-%d %H:%M:%S'),
            ip,
            location['country'],
            location['city'],
            location['region'],
            f"{user_agent.browser.family} {user_agent.browser.version_string}",
            f"{user_agent.os.family} {user_agent.os.version_string}",
            user_agent.device.family,
            data.get('referrer', 'Direct'),
            data.get('page_url', 'Unknown')
        ]

        # Save to Excel
        wb = load_workbook(EXCEL_FILE)
        ws = wb.active
        ws.append(visitor_data)
        wb.save(EXCEL_FILE)

        return JSONResponse({
            "status": "success",
            "message": "Visitor tracked successfully",
            "visitor_count": ws.max_row - 1
        })

    except Exception as e:
        print(f" Error: {str(e)}")
        return JSONResponse({"status": "error", "message": str(e)}, status_code=500)

@app.get("/stats")
def get_stats():
    try:
        wb = load_workbook(EXCEL_FILE)
        ws = wb.active
        return {"total_visitors": ws.max_row - 1}
    except:
        return {"total_visitors": 0}

# Initialize Excel when script starts
init_excel()
print("\n FastAPI Visitor Tracker Running!")

